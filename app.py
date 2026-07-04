from __future__ import annotations

import io
import re
import zipfile
from dataclasses import dataclass
from pathlib import Path
from typing import Dict, Iterable, List, Optional, Tuple

import openpyxl
import pandas as pd
import streamlit as st
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

STATUS_RE = re.compile(r"\b(DNC|PL|LD)\b|parental\s*leave|light\s*dut", re.I)
TRADE_RE = re.compile(
    r"\b(QCTO|CTO|CSA|STO|RTO|TTO|INACTIVE|SICK|LD|PL|DNC)\b.*$", re.I
)

DEFAULT_SCAN_COLUMNS = ["60 Hours", "Canvassing Day Shift", "Attempt 1", "Notes", "Canvassing Night Shift", "Attempt 2"]


@dataclass
class StatusHit:
    source_file: str
    sheet_name: str
    name_raw: str
    name_key: str
    source_row: int
    source_col: int
    source_header: str
    value: str


@dataclass
class UpdateResult:
    matched_name: str
    final_row: int
    final_col: int
    final_header: str
    old_value: str
    new_value: str
    source_file: str
    source_row: int


def clean_text(value) -> str:
    if value is None:
        return ""
    return str(value).strip()


def normalize_name(value: str) -> str:
    """Normalize crew names like 'Stone, Sean QCTO (LD)' to 'stone sean'."""
    text = clean_text(value)
    text = re.sub(r"\([^)]*\)", " ", text)
    text = TRADE_RE.sub("", text)
    text = re.sub(r"[^A-Za-z, '\-]", " ", text)
    text = re.sub(r"\s+", " ", text).strip().lower()
    if "," in text:
        last, first = [part.strip() for part in text.split(",", 1)]
        text = f"{last} {first}"
    text = text.replace("'", "").replace("-", " ")
    return re.sub(r"\s+", " ", text).strip()


def workbook_bytes(uploaded_file) -> bytes:
    if isinstance(uploaded_file, (bytes, bytearray)):
        return bytes(uploaded_file)
    if hasattr(uploaded_file, "getvalue"):
        return uploaded_file.getvalue()
    return Path(uploaded_file).read_bytes()


def load_workbook_from_bytes(data: bytes, keep_vba: bool = True):
    return openpyxl.load_workbook(io.BytesIO(data), keep_vba=keep_vba)


def iter_excel_uploads(files) -> Iterable[Tuple[str, bytes]]:
    """Yield (filename, bytes) from xlsx/xlsm uploads or zip uploads containing them."""
    for file in files or []:
        name = getattr(file, "name", "uploaded_file")
        data = workbook_bytes(file)
        suffix = Path(name).suffix.lower()
        if suffix == ".zip":
            with zipfile.ZipFile(io.BytesIO(data)) as zf:
                for info in sorted(zf.infolist(), key=lambda i: i.filename.lower()):
                    inner_suffix = Path(info.filename).suffix.lower()
                    if inner_suffix in {".xlsx", ".xlsm"} and not Path(info.filename).name.startswith("~$"):
                        yield Path(info.filename).name, zf.read(info)
        elif suffix in {".xlsx", ".xlsm"}:
            yield name, data


def row_values(ws: Worksheet, row: int) -> List[str]:
    return [clean_text(ws.cell(row, col).value) for col in range(1, ws.max_column + 1)]


def detect_header_row(ws: Worksheet) -> int:
    best_row = 1
    best_score = -1
    header_words = ["qcto", "ct", "name", "attempt", "canvassing", "notes", "predicted", "hours"]
    for row in range(1, min(ws.max_row, 20) + 1):
        values = " | ".join(row_values(ws, row)).lower()
        score = sum(1 for word in header_words if word in values)
        non_empty = sum(1 for value in row_values(ws, row) if value)
        score += min(non_empty, 8) / 10
        if score > best_score:
            best_row = row
            best_score = score
    return best_row


def detect_name_col(ws: Worksheet, header_row: int) -> int:
    # Prefer a header with QCTO/name/employee; otherwise pick the column with many comma names.
    headers = row_values(ws, header_row)
    for idx, header in enumerate(headers, start=1):
        h = header.lower()
        if h in {"name", "employee", "crew"} or "qcto" in h or "cto" == h:
            return idx
    best_col = 1
    best_score = -1
    for col in range(1, min(ws.max_column, 12) + 1):
        score = 0
        for row in range(header_row + 1, min(ws.max_row, header_row + 80) + 1):
            value = clean_text(ws.cell(row, col).value)
            if "," in value and len(normalize_name(value).split()) >= 2:
                score += 1
        if score > best_score:
            best_col = col
            best_score = score
    return best_col


def header_map(ws: Worksheet, header_row: int) -> Dict[int, str]:
    return {col: clean_text(ws.cell(header_row, col).value) or get_column_letter(col) for col in range(1, ws.max_column + 1)}


def collect_status_hits(source_files: Iterable[Tuple[str, bytes]], sheet_preference: Optional[str] = None) -> List[StatusHit]:
    hits: List[StatusHit] = []
    for filename, data in source_files:
        wb = load_workbook_from_bytes(data, keep_vba=True)
        sheet_names = wb.sheetnames
        chosen_sheets = []
        if sheet_preference and sheet_preference in sheet_names:
            chosen_sheets = [sheet_preference]
        else:
            chosen_sheets = [s for s in sheet_names if "source" not in s.lower()]
            if not chosen_sheets:
                chosen_sheets = [sheet_names[0]]
        for sheet_name in chosen_sheets:
            ws = wb[sheet_name]
            header_row = detect_header_row(ws)
            name_col = detect_name_col(ws, header_row)
            headers = header_map(ws, header_row)
            for row in range(header_row + 1, ws.max_row + 1):
                name_raw = clean_text(ws.cell(row, name_col).value)
                name_key = normalize_name(name_raw)
                if len(name_key.split()) < 2:
                    continue
                for col in range(1, ws.max_column + 1):
                    if col == name_col:
                        continue
                    value = clean_text(ws.cell(row, col).value)
                    if value and STATUS_RE.search(value):
                        hits.append(
                            StatusHit(
                                source_file=filename,
                                sheet_name=sheet_name,
                                name_raw=name_raw,
                                name_key=name_key,
                                source_row=row,
                                source_col=col,
                                source_header=headers.get(col, get_column_letter(col)),
                                value=value,
                            )
                        )
    return hits


def find_target_col(final_headers_by_col: Dict[int, str], source_col: int, source_header: str) -> int:
    # First match exact/non-case-sensitive header from source to final.
    norm_source = source_header.strip().lower()
    for col, header in final_headers_by_col.items():
        if header.strip().lower() == norm_source:
            return col

    # Common report variants: Attempt 1 ~= Canvassing Day Shift, Attempt 2 ~= Canvassing Night Shift.
    synonyms = {
        "attempt 1": ["canvassing day shift", "attempt 1"],
        "canvassing day shift": ["attempt 1", "canvassing day shift"],
        "attempt 2": ["canvassing night shift", "attempt 2"],
        "canvassing night shift": ["attempt 2", "canvassing night shift"],
        "hrs left": ["hours left", "hours left in 60"],
    }
    for candidate in synonyms.get(norm_source, []):
        for col, header in final_headers_by_col.items():
            if header.strip().lower() == candidate:
                return col

    # Fallback: keep same physical column if it exists.
    return source_col


def apply_updates(final_data: bytes, final_filename: str, hits: List[StatusHit], mode: str, sheet_name: Optional[str]) -> Tuple[bytes, List[UpdateResult], List[StatusHit]]:
    wb = load_workbook_from_bytes(final_data, keep_vba=Path(final_filename).suffix.lower() == ".xlsm")
    ws = wb[sheet_name] if sheet_name and sheet_name in wb.sheetnames else wb[wb.sheetnames[0]]
    header_row = detect_header_row(ws)
    name_col = detect_name_col(ws, header_row)
    headers = header_map(ws, header_row)

    final_names: Dict[str, int] = {}
    duplicates = set()
    for row in range(header_row + 1, ws.max_row + 1):
        key = normalize_name(clean_text(ws.cell(row, name_col).value))
        if len(key.split()) < 2:
            continue
        if key in final_names:
            duplicates.add(key)
        else:
            final_names[key] = row

    updates: List[UpdateResult] = []
    unmatched: List[StatusHit] = []

    # Later files/hits should win, but we still record each update. The uploaded ZIP order is sorted by filename.
    for hit in hits:
        final_row = final_names.get(hit.name_key)
        if not final_row or hit.name_key in duplicates:
            unmatched.append(hit)
            continue
        target_col = find_target_col(headers, hit.source_col, hit.source_header)
        old_value = clean_text(ws.cell(final_row, target_col).value)
        new_value = hit.value
        if mode == "Append to existing cell" and old_value and old_value.lower() != new_value.lower():
            if new_value.lower() not in old_value.lower():
                new_value = f"{old_value}; {new_value}"
            else:
                new_value = old_value
        ws.cell(final_row, target_col).value = new_value
        updates.append(
            UpdateResult(
                matched_name=hit.name_raw,
                final_row=final_row,
                final_col=target_col,
                final_header=headers.get(target_col, get_column_letter(target_col)),
                old_value=old_value,
                new_value=new_value,
                source_file=hit.source_file,
                source_row=hit.source_row,
            )
        )

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue(), updates, unmatched


def df_hits(hits: List[StatusHit]) -> pd.DataFrame:
    return pd.DataFrame([
        {
            "source_file": h.source_file,
            "sheet": h.sheet_name,
            "name": h.name_raw,
            "normalized_name": h.name_key,
            "source_row": h.source_row,
            "source_column": get_column_letter(h.source_col),
            "source_header": h.source_header,
            "value": h.value,
        }
        for h in hits
    ])


def df_updates(updates: List[UpdateResult]) -> pd.DataFrame:
    return pd.DataFrame([
        {
            "name": u.matched_name,
            "final_row": u.final_row,
            "final_column": get_column_letter(u.final_col),
            "final_header": u.final_header,
            "old_value": u.old_value,
            "new_value": u.new_value,
            "source_file": u.source_file,
            "source_row": u.source_row,
        }
        for u in updates
    ])


st.set_page_config(page_title="Crew DNC / PL / LD Updater", layout="wide")
st.title("Crew DNC / PL / LD Updater")
st.caption("Upload prior canvassing reports/lists, then upload the final report to mark matching crew as DNC, PL, or LD.")

with st.expander("How this app matches and updates", expanded=True):
    st.markdown(
        """
- Reads `.xlsx` and `.xlsm` files directly, plus `.zip` files containing those workbooks.
- Finds names like `Last, First QCTO` and normalizes them to match across files.
- Scans uploaded list/report rows for `DNC`, `PL`, `LD`, `Parental Leave`, or `Light Duties`.
- Updates the matching person in the final workbook, preferably in the same header column. If headers differ, it uses the same physical column.
- Download preserves the final workbook format as much as OpenPyXL allows. For `.xlsm`, macros are kept when possible.
        """
    )

left, right = st.columns(2)
with left:
    final_file = st.file_uploader("1) Final/master report to update", type=["xlsx", "xlsm"])
with right:
    source_files = st.file_uploader(
        "2) DNC / PL / LD list files or ZIP",
        type=["xlsx", "xlsm", "zip"],
        accept_multiple_files=True,
    )

mode = st.radio("If the destination cell already has text", ["Overwrite cell", "Append to existing cell"], horizontal=True)

if source_files:
    source_payloads = list(iter_excel_uploads(source_files))
    hits = collect_status_hits(source_payloads)
    st.subheader("Detected DNC / PL / LD entries")
    c1, c2, c3 = st.columns(3)
    c1.metric("Workbooks read", len(source_payloads))
    c2.metric("Status entries found", len(hits))
    c3.metric("Unique people", len({h.name_key for h in hits}))
    if hits:
        st.dataframe(df_hits(hits), use_container_width=True, hide_index=True)
    else:
        st.warning("No DNC / PL / LD entries were found in the uploaded list files.")
else:
    hits = []

if final_file and hits:
    final_data = workbook_bytes(final_file)
    final_wb = load_workbook_from_bytes(final_data, keep_vba=Path(final_file.name).suffix.lower() == ".xlsm")
    sheet_name = st.selectbox("Final workbook sheet to update", final_wb.sheetnames, index=0)

    if st.button("Update final workbook", type="primary"):
        updated_bytes, updates, unmatched = apply_updates(final_data, final_file.name, hits, mode, sheet_name)
        st.success(f"Updated {len(updates)} cells. {len(unmatched)} detected status entries were unmatched or duplicate names.")

        col_a, col_b = st.columns(2)
        with col_a:
            st.subheader("Updates made")
            st.dataframe(df_updates(updates), use_container_width=True, hide_index=True)
        with col_b:
            st.subheader("Unmatched / skipped")
            st.dataframe(df_hits(unmatched), use_container_width=True, hide_index=True)

        suffix = Path(final_file.name).suffix or ".xlsx"
        output_name = f"{Path(final_file.name).stem}_DNC_PL_LD_updated{suffix}"
        st.download_button(
            "Download updated workbook",
            data=updated_bytes,
            file_name=output_name,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
elif final_file and not hits:
    st.info("Upload DNC / PL / LD source lists first, then the update button will appear.")
elif source_files and not final_file:
    st.info("Now upload the final/master workbook that should be updated.")
else:
    st.info("Start by uploading your final report and the 5 DNC / PL / LD list files or ZIP.")
